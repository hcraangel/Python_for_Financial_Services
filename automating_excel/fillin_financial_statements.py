from datetime import date, datetime
from openpyxl import Workbook, load_workbook
import os

def choose_fill_func(release_date, expiration_date):
    #workbook filling funcions
    def fill_func_1(templates, interest, principle):
        wb = load_workbook(templates[0])
        wb["balance_sheet"]['C9'].value = principle
        wb["balance_sheet"]['B32'].value = interest[0]
        wb["balance_sheet"]['C13'].value = interest[1]
        wb["income_statement"]['C5'].value = interest[3]
        wb["income_statement"]['B5'].value = interest[4]
        return wb
    
    def fill_func_2(templates, interest, principle):
        wb = load_workbook(templates[1])
        wb["balance_sheet"]['B9'].value = principle
        wb["income_statement"]['B32'].value = interest[0]
        wb["balance_sheet"]['C13'].value = interest[1]
        wb["balance_sheet"]['B13'].value = interest[2]
        wb["income_statement"]['C5'].value = interest[3]
        wb["income_statement"]['B5'].value = interest[4]
        return wb
    
    def fill_func_3(templates, interest, principle):
        wb = load_workbook(templates[2])
        wb["statement_of_cash_flows"]['B18'].value = principle
        wb["income_statement"]['B32'].value = interest[0]
        wb["income_statement"]['B5'].value = interest[2]
        wb["balance_sheet"]['C5'].value = interest[5]
        return wb
    
    def fill_func_4(templates, interest, principle):
        wb = load_workbook(templates[3])
        wb["balance_sheet"]['B9'].value = principle
        wb["balance_sheet"]['B13'].value = interest[2]
        return wb
    
    def fill_func_5(templates, interest, principle):
        wb = load_workbook(templates[4])
        wb["balance_sheet"]['B9'].value = principle
        wb["balance_sheet"]['B5'].value = interest[4]
        wb["balance_sheet"]['B32'].value = interest[0]
        wb["balance_sheet"]['B13'].value = interest[2]
        return wb
    
    def fill_func_void(templates, interest, principle):
        pass
    
    
    released_last_year = release_date < end_of_last_year
    released_this_year_before_this_month = release_date > end_of_last_year and release_date <= end_of_last_month
    released_this_month = release_date > end_of_last_month and release_date <= end_of_this_month
    released_after_this_month = release_date > end_of_this_month
    matured_this_month = expiration_date <= end_of_this_month
    
    if released_last_year and matured_this_month:
        return fill_func_1
    elif released_last_year and not matured_this_month:
        return fill_func_2
    elif released_this_month and matured_this_month:
        return fill_func_3
    elif released_this_month and not matured_this_month:
        return fill_func_4
    elif released_this_year_before_this_month and not matured_this_month:
        return fill_func_5
    elif released_after_this_month:
        print("time frame out of range")
        return fill_func_void
    else:
        print("unpredicted condition")
        return fill_func_void

#end of choose_fill_func

source_wb = load_workbook('./source_workbook.xlsx')
source_ws = source_wb["sheet1"]

rowLen = source_ws.max_row
colLen = source_ws.max_column

end_of_last_year = date(2016, 12, 31)
end_of_last_month = date(2017, 4, 30)
end_of_this_month = date(2017, 5, 31)

os.makedirs('./results')
result_dir = './results'

for row in source_ws.iter_rows(min_row=2, max_row=227, max_col=colLen):
    release_date = row[3].value.date()
    expiration_date = row[4].value.date()
    principle = row[1].value
    rate_of_return = row[2].value
    
    interest = [
                (end_of_last_month - release_date).days*principle*rate_of_return/365,
                (end_of_last_year - release_date).days*principle*rate_of_return/365,
                (end_of_this_month - release_date).days*principle*rate_of_return/365,
                (end_of_this_month - end_of_last_year).days*principle*rate_of_return/365,
                (end_of_this_month - end_of_last_month).days*principle*rate_of_return/365,
                (expiration_date - release_date).days*principle*rate_of_return/365
                ]
        
    template_path = ['./template1.xlsx','./template2.xlsx','./template3.xlsx','./template4.xlsx','./template5.xlsx']
                
    target_path = result_dir + '/' + str(row[0].value) + '.xlsx'
    fill_func = choose_fill_func(release_date, expiration_date)
    result_spreadsheet = fill_func(template_path, interest, principle)
    if result_spreadsheet is not None:
        result_spreadsheet.save(target_path)









